#!/usr/bin/env python

import pandas as pd
from glob import glob
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Define paths and dataset mapping
model_result_paths = glob("./data/judgements/*/*/*.json")

judge_dict = {
    "judge_gpt-4-turbo-preview": "GPT4", # "gpt-4-0125-preview"
    "judge_llmjudge-athenev2": "AtheneV2",
    "judge_llmjudge-llama33": "Llama 3.3",
    "judge_llmjudge-tulu405": "Tulu405",
}

eval_dataset_dict = {
    "elyza__ELYZA-tasks-100": "ELYZA 100",
    "yuzuai__rakuda-questions": "Rakuda",
    "lightblue__tengu_bench": "Tengu",
    "shisa-ai__ja-mt-bench-1shot": "JA-MT",
}

# Initialize dictionaries to store results by judge
judge_results = defaultdict(list)
model_judge_counts = defaultdict(set)

# Process results by judge
for model_result_path in model_result_paths:
    parts = model_result_path.split("/")
    judge_id = parts[3]
    eval_dataset_key = parts[4]
    model_name = parts[5].replace(".json", "")
    
    # Skip if not a known judge
    if judge_id not in judge_dict:
        continue
    
    # Track which judges have scored each model
    model_judge_counts[model_name].add(judge_id)
        
    df = pd.read_json(model_result_path, lines=True)
    df["eval_dataset"] = eval_dataset_dict[eval_dataset_key]
    df["model_name"] = model_name
    
    # Apply ELYZA score adjustment (0-5 to 0-10 scale)
    if eval_dataset_dict[eval_dataset_key] == "ELYZA 100":
        df["score"] = df["score"] * 2
    
    judge_results[judge_id].append(df)

# Process results for each judge and combine
all_judge_means = {}
all_dfs = []

# Process each judge's results
for judge_id, dfs in judge_results.items():
    if not dfs:  # Skip if no results for this judge
        continue
        
    judge_df = pd.concat(dfs)
    judge_means = judge_df.groupby(["model_name", "eval_dataset"]).score.mean().unstack()
    
    # Add average and rename columns to add judge name
    judge_means.columns = [f"{col} ({judge_dict[judge_id]})" for col in judge_means.columns]
    judge_means[f"Average ({judge_dict[judge_id]})"] = judge_means.mean(axis=1)
    
    all_judge_means[judge_id] = judge_means
    all_dfs.append(judge_df)

# Calculate overall averages across all judges
if all_dfs:
    all_result_df = pd.concat(all_dfs)
    overall_means = all_result_df.groupby(["model_name", "eval_dataset"]).score.mean().unstack()
    overall_means["Average (All)"] = overall_means.mean(axis=1)
    overall_means.columns = [f"{col} (All)" if col != "Average (All)" else col for col in overall_means.columns]

    # Combine all results
    result_df = overall_means
    for judge_means in all_judge_means.values():
        result_df = result_df.join(judge_means, how='outer')
    
    # Add judge count column
    result_df["# Judges"] = pd.Series({model: len(judges) for model, judges in model_judge_counts.items()})

    # Sort by Average (All) descending
    result_df = result_df.sort_values("Average (All)", ascending=False)

    # Replace '__' with '/' in model names (index)
    result_df.index = result_df.index.str.replace('__', '/')

    # Calculate Average (ex-GPT4)
    non_gpt4_averages = [col for col in result_df.columns if col.startswith("Average") and "GPT4" not in col]
    result_df["Average (ex-GPT4)"] = result_df[non_gpt4_averages].mean(axis=1)
    # Set to NaN if only GPT4 scores are available
    gpt4_only_mask = result_df[[col for col in non_gpt4_averages if "(All)" not in col]].isna().all(axis=1)
    result_df.loc[gpt4_only_mask, "Average (ex-GPT4)"] = pd.NA

    # Calculate GPT4 delta (percentage difference from GPT4 baseline)
    gpt4_avg = result_df["Average (GPT4)"].astype('float64')
    ex_gpt4_avg = result_df["Average (ex-GPT4)"].astype('float64')
    # Calculate percentage difference where both scores exist
    result_df["GPT4 delta"] = pd.NA  # Initialize with NaN
    valid_scores = ~(gpt4_avg.isna() | ex_gpt4_avg.isna())
    
    # Debug print some values
    # Ensure floating point division
    delta = (ex_gpt4_avg - gpt4_avg) / gpt4_avg
    print("Raw delta:", delta[valid_scores].head())
    
    result_df.loc[valid_scores, "GPT4 delta"] = delta[valid_scores]

    # Organize columns by judge blocks
    # First, get the 'All' block
    all_cols = ["Average (All)"] + [col for col in result_df.columns if col.endswith("(All)") and col != "Average (All)"]
    
    # Then get each judge's block
    judge_blocks = []
    for judge_name in sorted(judge_dict.values()):
        judge_cols = [col for col in result_df.columns if f"({judge_name})" in col]
        # Sort within judge block to have Average first
        avg_col = [col for col in judge_cols if col.startswith("Average")]
        other_cols = sorted([col for col in judge_cols if not col.startswith("Average")])
        judge_blocks.extend(avg_col + other_cols)
    
    # Combine all blocks in order, with # Judges and Average (ex-GPT4) after index
    sorted_cols = ["# Judges", "Average (ex-GPT4)"] + all_cols + judge_blocks + ["GPT4 delta"]
    result_df = result_df[sorted_cols]

    # Create Excel writer with formatting
    output_file = 'output-multi.xlsx'
    
    # First save to Excel using pandas to get the basic structure
    # Export to Excel with more precision for delta column
    result_df.to_excel(output_file, float_format='%.4f', index_label="Model")
    
    # Then apply additional formatting using openpyxl
    from openpyxl import load_workbook
    wb = load_workbook(output_file)
    ws = wb.active
    
    # Define styles
    header_font = Font(name='Source Code Pro', bold=True)
    cell_font = Font(name='Source Code Pro')
    left_align = Alignment(horizontal='left', vertical='top')
    header_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    number_align = Alignment(horizontal='right', vertical='top')
    
    # Apply formatting to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = cell_font
            cell.border = None
            
            if cell.row == 1:  # Header row
                cell.font = header_font
                cell.fill = PatternFill(start_color="dde8cb", end_color="dde8cb", fill_type="solid")
                cell.alignment = header_align
            elif cell.column == 1:  # Model column
                cell.font = header_font
                cell.alignment = left_align
                # Color entire row if model name contains 'shisa'
                if 'shisa' in str(cell.value).lower():
                    for row_cell in row:
                        row_cell.fill = PatternFill(start_color="dcb1ff", end_color="dcb1ff", fill_type="solid")
            else:  # Numeric cells
                cell.alignment = number_align
                # Set specific number format for GPT4 delta column
                if ws.cell(1, cell.column).value == "GPT4 delta":
                    cell.number_format = '0.0%'
    
    # Freeze panes for first row and column
    ws.freeze_panes = 'B2'
    
    # Set specific column widths
    # Model column (A)
    ws.column_dimensions['A'].width = 60
    
    # # Judges column (B)
    ws.column_dimensions['B'].width = 12
    
    # All other numeric columns
    for column in range(3, ws.max_column + 1):
        col_letter = get_column_letter(column)
        ws.column_dimensions[col_letter].width = 14
    
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill, Font
    
    # Get column indices for score columns (skip Model, # Judges, and GPT4 delta)
    score_columns = list(range(3, ws.max_column))  # Columns C onwards, excluding last column
    
    # Add conditional formatting for each score column
    for col in score_columns:
        col_letter = get_column_letter(col)
        # Range for the column, excluding header
        col_range = f"{col_letter}2:{col_letter}{ws.max_row}"
        
        # Create rule to highlight maximum value
        # Use absolute reference for MAX range but relative for current cell
        rule = CellIsRule(
            operator='equal',
            formula=[f'MAX(${col_letter}$2:${col_letter}${ws.max_row})'],
            stopIfTrue=True,
            font=Font(name='Source Code Pro', bold=True, color='127622')
        )
        
        # Apply the rule to the column range
        ws.conditional_formatting.add(col_range, rule)
    
    # Format numbers
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if cell.column == ws.max_column:  # GPT4 delta column
                    cell.number_format = '+0.0%;-0.0%;0%'
                else:
                    cell.number_format = '0.00'
    
    # Add threshold borders
    from openpyxl.styles import Border, Side
    threshold_border = Border(top=Side(style='thin'))
    
    # Find column indices for specific columns
    header_row = list(ws.iter_rows(min_row=1, max_row=1))[0]
    elyza_col = None
    jamt_col = None
    for cell in header_row:
        if cell.value == "ELYZA 100 (All)":
            elyza_col = cell.column
        elif cell.value == "JA-MT (All)":
            jamt_col = cell.column
    
    # Add borders at thresholds
    if elyza_col:
        # Find first value <= 8.2 in ELYZA column
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=elyza_col)
            if isinstance(cell.value, (int, float)) and cell.value <= 8.2:
                cell.border = threshold_border
                break
    
    if jamt_col:
        # Find first value <= 8.1 in JA-MT column
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=jamt_col)
            if isinstance(cell.value, (int, float)) and cell.value <= 8.1:
                cell.border = threshold_border
                break
    
    # Save the formatted workbook
    wb.save(output_file)
    print(f"Results saved to {output_file}")
else:
    print("No results found for any judge")